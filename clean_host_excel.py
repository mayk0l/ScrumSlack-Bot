import openpyxl
import shutil
import os

# Copy from template
shutil.copy("excel/Bitacora-Rentabilidad-Valuelist.xlsx", "project_tracking.xlsx")

wb = openpyxl.load_workbook("project_tracking.xlsx")

# 1. Clean Planificación
ws = wb["Planificación"]
max_r = ws.max_row
if max_r >= 2:
    ws.delete_rows(2, max_r) # delete_rows(idx, amount) - this deletes ALL rows from 2 onwards

# 2. Clean Administración
ws = wb["Administración"]
max_r = ws.max_row
if max_r >= 2:
    ws.delete_rows(2, max_r)

# 3. Clean Evidencias
if "Evidencias" in wb.sheetnames:
    ws = wb["Evidencias"]
    max_r = ws.max_row
    if max_r >= 2:
        ws.delete_rows(2, max_r)

# 4. Clean Bitácora
ws = wb["Bitácora"]
mapping = {"OG": 2, "OE1": 3, "OE2": 4, "OE3": 5, "OE4": 6, "OE5": 7, "OE6": 8}
for obj_id, r in mapping.items():
    ws.cell(row=r, column=2, value="OG" if obj_id == "OG" else obj_id) # Set cell B
    ws.cell(row=r, column=3, value="") # Clear cell C

# 5. Clean Gantt page (clear yellow boxes if it's conditional formatting or data)
# Wait, Gantt in Excel might have formulas or data. We shouldn't delete rows, just the data it reads.
# The user says "la página gantt aun tiene cuadros amarillos".
# The Gantt is usually drawn based on the "Planificación" sheet. If Planificación is empty, Gantt shouldn't have yellow boxes unless there's hardcoded data.
if "Gantt" in wb.sheetnames:
    ws = wb["Gantt"]
    # Often, Gantt charts have hardcoded task names on the left if not linked via formula
    for row in ws.iter_rows(min_row=5, max_row=50, min_col=1, max_col=1):
        if row[0].value:
            row[0].value = None

wb.save("project_tracking.xlsx")
print("Cleaned project_tracking.xlsx correctly.")
