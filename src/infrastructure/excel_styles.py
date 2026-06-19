"""Módulo: excel_styles.

Primitivos de estilo reutilizables para la generación de archivos Excel:
paleta corporativa, headers, bordes, tablas nativas y colores tipo semáforo.
Centralizar estos estilos evita duplicación entre hojas y servicios y asegura
una presentación consistente en todo el archivo `project_tracking.xlsx`.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

# --- Paleta corporativa ---
COLOR_HEADER = "1F4E78"      # Azul oscuro (headers)
COLOR_SUBHEADER = "D9E1F2"   # Azul claro (sub-headers / timeline Gantt)
COLOR_DATABAR = "5A8AC6"     # Azul (barras de progreso DataBar)
COLOR_GANTT_BAR = "4472C4"   # Azul (barras de tareas en Gantt)
COLOR_OG = "E7E6E6"          # Gris (Objetivo General)
COLOR_OE = "F2F2F2"          # Gris claro (Objetivos Específicos)
COLOR_WHITE = "FFFFFF"

# --- Semáforo por severidad de riesgo ---
SEVERITY_COLORS = {
    "low": "A9D08E",       # verde
    "medium": "FFD966",    # amarillo
    "high": "F4B084",      # naranja
    "critical": "FF6B6B",  # rojo
}

# --- Semáforo por estado de tarea ---
STATUS_COLORS = {
    "NO COMENZADO": "D9D9D9",  # gris
    "EN CURSO": "FFD966",      # amarillo
    "COMPLETADO": "A9D08E",    # verde
    "BLOQUEADO": "FF6B6B",     # rojo
}

TABLE_STYLE = "TableStyleMedium9"


def solid_fill(color: str) -> PatternFill:
    """Relleno sólido a partir de un color hexadecimal."""
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def header_fill() -> PatternFill:
    return solid_fill(COLOR_HEADER)


def header_font() -> Font:
    return Font(color=COLOR_WHITE, bold=True)


def thin_border() -> Border:
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def align_center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def align_left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws: Worksheet, max_col: int | None = None) -> None:
    """Aplica el estilo corporativo (azul, negrita, blanco) a la fila 1."""
    fill = header_fill()
    font = header_font()
    alignment = align_center()
    border = thin_border()
    cols = max_col or ws.max_column
    for col in range(1, cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border


def ensure_table(ws: Worksheet, name: str, ref: str) -> None:
    """Crea o actualiza una tabla nativa con filtros y zebra stripes."""
    if name in ws.tables:
        ws.tables[name].ref = ref
        return
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
