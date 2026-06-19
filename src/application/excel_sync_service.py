"""Módulo: excel_sync_service.

Servicio de aplicación para sincronización con archivo Excel.
"""

from __future__ import annotations

import asyncio
from pathlib import Path
from typing import Any
from uuid import UUID

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.domain.models import MetricSnapshot, Risk
from src.domain.repositories import (
    MetricRepository,
    RiskRepository,
    SprintRepository,
)

SHEET_MODULES = "Módulos"
SHEET_GANTT = "Carta Gantt"
SHEET_TASKS = "Tareas"
SHEET_METRICS = "Métricas"
SHEET_RISKS = "Riesgos"


class ExcelSyncService:
    """Crea y actualiza el archivo Excel de seguimiento."""

    def __init__(
        self,
        excel_path: str,
        metric_repo: MetricRepository,
        risk_repo: RiskRepository,
        sprint_repo: SprintRepository,
    ):
        self._excel_path = Path(excel_path)
        self._metric_repo = metric_repo
        self._risk_repo = risk_repo
        self._sprint_repo = sprint_repo

    async def create_template(self) -> None:
        """Crea el archivo Excel con las 5 hojas y headers."""

        def _create() -> None:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            self._setup_sheet(wb, SHEET_MODULES, ["ID", "Nombre", "Estado", "Progreso %"])
            self._setup_sheet(wb, SHEET_GANTT, ["Sprint", "Inicio", "Fin", "Estado"])
            self._setup_sheet(
                wb, SHEET_TASKS, ["ID", "Sprint", "Responsable", "Estado", "Descripción"]
            )
            self._setup_sheet(
                wb, SHEET_METRICS, ["Fecha", "Tipo", "Valor", "Sprint", "Metadata"]
            )
            self._setup_sheet(
                wb, SHEET_RISKS, ["Fecha", "Tipo", "Severidad", "Descripción", "Resuelto"]
            )

            self._excel_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(self._excel_path)

        await asyncio.to_thread(_create)

    def _setup_sheet(
        self, wb: Workbook, title: str, headers: list[str]
    ) -> Worksheet:
        """Crea una hoja con headers y aplica estilo corporativo pro."""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        ws = wb.create_sheet(title=title)
        ws.append(headers)
        
        # Estilos premium
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align
            cell.border = thin_border
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = max(len(str(cell.value)) + 8, 18)
        
        # Tabla nativa con filtros y zebra stripes
        table_name = f"Tabla_{title.replace(' ', '')[:20]}"
        tab = Table(displayName=table_name, ref=f"A1:{get_column_letter(len(headers))}2")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
            
        ws.freeze_panes = "A2"
        return ws

    async def sync_metrics(self, team_id: UUID) -> None:
        """Escribe las métricas del equipo en la hoja Métricas."""
        metric = await self._metric_repo.get_latest(team_id, "velocity")
        metrics = [metric] if metric else []

        def _sync(data: list[MetricSnapshot]) -> None:
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.styles import Alignment, numbers
            
            wb = openpyxl.load_workbook(self._excel_path)
            ws = wb[SHEET_METRICS]
            
            # Eliminar filas manteniendo tabla
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            for item in data:
                ws.append([
                    item.date.isoformat(),
                    item.metric_type,
                    item.value,
                    item.metadata.get("sprint_id", ""),
                    str(item.metadata),
                ])
            
            # Actualizar rango de tabla
            max_r = max(2, ws.max_row)
            table_name = "Tabla_Métricas"
            if table_name in ws.tables:
                ws.tables[table_name].ref = f"A1:E{max_r}"
            
            # Formatear celdas
            for row in ws.iter_rows(min_row=2, max_row=max_r):
                row[0].number_format = numbers.FORMAT_DATE_XLSX14
                row[2].number_format = numbers.FORMAT_NUMBER_00
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
            wb.save(self._excel_path)

        await asyncio.to_thread(_sync, metrics)

    async def sync_risks(self, team_id: UUID) -> None:
        """Escribe los riesgos activos en la hoja Riesgos."""
        risks = await self._risk_repo.get_active_by_team(team_id)

        def _sync(data: list[Risk]) -> None:
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.styles import Alignment, PatternFill, Font, numbers
            
            wb = openpyxl.load_workbook(self._excel_path)
            ws = wb[SHEET_RISKS]
            
            # Eliminar filas manteniendo tabla
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            severity_colors = {
                "low": PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid"),
                "medium": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
                "high": PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"),
                "critical": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            }

            for risk in data:
                row_idx = ws.max_row + 1
                ws.append([
                    risk.created_at.isoformat(),
                    risk.type.value,
                    risk.severity.value.upper(),
                    risk.description,
                    "✅ Sí" if risk.resolved else "⚠️ No",
                ])
                
                # Aplicar color por severidad
                severity_cell = ws.cell(row=row_idx, column=3)
                if risk.severity.value in severity_colors:
                    severity_cell.fill = severity_colors[risk.severity.value]
                    severity_cell.font = Font(bold=True)
            
            # Actualizar rango de tabla
            max_r = max(2, ws.max_row)
            table_name = "Tabla_Riesgos"
            if table_name in ws.tables:
                ws.tables[table_name].ref = f"A1:E{max_r}"
            
            # Formatear celdas
            for row in ws.iter_rows(min_row=2, max_row=max_r):
                row[0].number_format = numbers.FORMAT_DATE_XLSX14
                for cell in row:
                    if cell.column != 4:  # No centrar descripción
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        
            wb.save(self._excel_path)

        await asyncio.to_thread(_sync, risks)

    async def get_module_progress(self) -> list[dict[str, Any]]:
        """Lee la hoja Módulos y retorna progreso."""

        def _read() -> list[dict]:
            import zipfile
            try:
                wb = openpyxl.load_workbook(self._excel_path)
            except (FileNotFoundError, zipfile.BadZipFile):
                return []
            if "Módulos" not in wb.sheetnames:
                return []
            ws = wb[SHEET_MODULES]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                rows.append(
                    {
                        "id": row[0],
                        "name": row[1],
                        "status": row[2],
                        "progress": row[3],
                    }
                )
            return rows

        return await asyncio.to_thread(_read)

    async def update_module(self, module_id: str, updates: dict[str, Any]) -> None:
        """Actualiza un módulo específico en la hoja Módulos."""

        def _update() -> None:
            wb = openpyxl.load_workbook(self._excel_path)
            ws = wb[SHEET_MODULES]
            for idx, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                if row[0] == module_id:
                    if "name" in updates:
                        ws.cell(row=idx, column=2, value=updates["name"])
                    if "status" in updates:
                        ws.cell(row=idx, column=3, value=updates["status"])
                    if "progress" in updates:
                        ws.cell(row=idx, column=4, value=updates["progress"])
                    break
            wb.save(self._excel_path)

        await asyncio.to_thread(_update)
