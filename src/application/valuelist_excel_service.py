import asyncio
import re
import openpyxl
from typing import Any

# --- Esquema canónico de las hojas de tareas (Planificación / Administración) ---
TASK_HEADERS = [
    "Actividad",
    "Descripción (resumen)",
    "Responsable",
    "Comienzo",
    "Fin (Esperado/logrado)",
    "Estado",
    "% de logro",
    "Entregable",
    "Comentarios",
]

# Índices 0-based para iter_rows(values_only=True).
(
    IDX_ACTIVIDAD,
    IDX_DESC,
    IDX_RESP,
    IDX_INICIO,
    IDX_FIN,
    IDX_ESTADO,
    IDX_PROGRESO,
    IDX_ENTREGABLE,
    IDX_COMENT,
) = range(9)

# Estados válidos de una tarea (semáforo).
STATUS_NOT_STARTED = "NO COMENZADO"
STATUS_IN_PROGRESS = "EN CURSO"
STATUS_DONE = "COMPLETADO"
STATUS_BLOCKED = "BLOQUEADO"
TASK_STATUSES = [STATUS_NOT_STARTED, STATUS_IN_PROGRESS, STATUS_DONE, STATUS_BLOCKED]


def to_fraction(value: Any) -> float:
    """Normaliza un porcentaje a fracción 0-1.

    Acepta tanto fracciones (0.5) como porcentajes (50) y recorta al rango
    [0, 1], de modo que la lectura sea robusta sin importar cómo el usuario
    haya escrito el valor en el Excel.
    """
    try:
        v = float(str(value).replace("%", "").strip())
    except (TypeError, ValueError):
        return 0.0
    if v > 1:
        v = v / 100.0
    return max(0.0, min(1.0, v))


def derive_status(progress: float, current: str = "") -> str:
    """Deriva el Estado a partir del progreso, respetando BLOQUEADO manual."""
    current_norm = str(current).strip().upper() if current else ""
    if progress >= 1.0:
        return STATUS_DONE
    if current_norm == STATUS_BLOCKED:
        return STATUS_BLOCKED
    if progress > 0:
        return STATUS_IN_PROGRESS
    return STATUS_NOT_STARTED


class ValuelistExcelService:
    """
    Servicio para sincronizar y leer datos directamente desde 
    la Bitácora de Rentabilidad de Valuelist (Excel).
    """

    def __init__(self, excel_path: str):
        self._excel_path = excel_path

    async def validate_excel_file(self, file_content: bytes) -> tuple[bool, str]:
        """Valida que el contenido de un archivo Excel sea compatible con el sistema.
        
        Retorna (True, "") si es válido, o (False, "mensaje de error") si falla la validación.
        """
        import io
        from datetime import datetime, date
        
        def _validate() -> tuple[bool, str]:
            try:
                # 1. Cargar el libro en memoria
                wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            except Exception as e:
                return False, f"El archivo no es un libro de Excel válido (.xlsx) o está corrupto. Error: {str(e)}"
            
            # 2. Verificar hojas requeridas
            required_sheets = ["Bitácora", "Planificación", "Administración", "Evidencia"]
            for s in required_sheets:
                if s not in wb.sheetnames:
                    return False, f"Falta la pestaña obligatoria: '{s}'."
            
            # 3. Validar hoja Bitácora
            ws_bit = wb["Bitácora"]
            if ws_bit.max_column < 3:
                return False, "La pestaña 'Bitácora' debe tener al menos 3 columnas (Identificador, Campo, Valor/Descripción)."
                
            # 4. Validar hojas de tareas (Planificación y Administración)
            for sheet_name in ["Planificación", "Administración"]:
                ws = wb[sheet_name]
                if ws.max_column < 7:
                    return False, f"La pestaña '{sheet_name}' debe tener al menos 7 columnas (ID, Descripción, Responsable, Inicio, Fin, Estado, % Logro)."
                
                # Validar unicidad de IDs de tareas y tipos de datos por fila
                seen_ids = set()
                row_idx = 1
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_idx += 1
                    act_id = row[0]
                    if not act_id:
                        continue
                    
                    act_str = str(act_id).strip()
                    if not (act_str.startswith("A") or act_str.startswith("AD")):
                        continue
                        
                    # Validar IDs duplicados
                    if act_str in seen_ids:
                        return False, f"ID de tarea duplicado '{act_str}' en la pestaña '{sheet_name}', fila {row_idx}."
                    seen_ids.add(act_str)
                    
                    # Validar Responsable
                    resp = row[2]
                    if not resp or not str(resp).strip():
                        return False, f"Falta el 'Responsable' para la tarea '{act_str}' en la pestaña '{sheet_name}', fila {row_idx}."
                        
                    # Validar Fechas (columnas index 3 y 4)
                    for col_name, col_idx in [("Comienzo", 3), ("Fin", 4)]:
                        val = row[col_idx]
                        if val is not None:
                            if not isinstance(val, (datetime, date)):
                                try:
                                    datetime.strptime(str(val).strip()[:10], "%Y-%m-%d")
                                except ValueError:
                                    return False, f"Formato de fecha inválido en la columna '{col_name}' para la tarea '{act_str}' en la pestaña '{sheet_name}', fila {row_idx}. Debe ser YYYY-MM-DD o formato de fecha de Excel."
                    
                    # Validar Estado (columna index 5)
                    estado_val = row[IDX_ESTADO]
                    if estado_val is not None and str(estado_val).strip():
                        if str(estado_val).strip().upper() not in TASK_STATUSES:
                            return False, f"Estado inválido '{estado_val}' para la tarea '{act_str}' en la pestaña '{sheet_name}', fila {row_idx}. Debe ser uno de: {', '.join(TASK_STATUSES)}."

                    # Validar % logro (columna index 6)
                    logro = row[6]
                    if logro is not None:
                        try:
                            val_logro = float(logro)
                            if val_logro < 0.0 or val_logro > 100.0:
                                return False, f"El porcentaje de logro para la tarea '{act_str}' en la pestaña '{sheet_name}', fila {row_idx} debe estar entre 0 y 100 (o 0.0 y 1.0). Valor recibido: {logro}."
                        except ValueError:
                            return False, f"El porcentaje de logro para la tarea '{act_str}' en la pestaña '{sheet_name}', fila {row_idx} debe ser un valor numérico. Recibido: '{logro}'."
            
            return True, ""
            
        return await asyncio.to_thread(_validate)

    @staticmethod
    def _apply_gantt_and_styles(wb: openpyxl.Workbook):
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.formatting.rule import DataBarRule, CellIsRule
        from openpyxl.worksheet.datavalidation import DataValidation
        from datetime import datetime, timedelta
        from src.infrastructure import excel_styles as xls

        header_fill = xls.header_fill()
        header_font = xls.header_font()
        align_center = xls.align_center()
        align_left = xls.align_left()
        thin_border = xls.thin_border()

        headers = TASK_HEADERS
        
        # 0. Format Bitácora (Executive Summary)
        if "Bitácora" in wb.sheetnames:
            ws_bit = wb["Bitácora"]
            ws_bit.freeze_panes = "A2"
            
            # Style headers
            for row in ws_bit.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
            
            # Column widths
            ws_bit.column_dimensions['A'].width = 8
            ws_bit.column_dimensions['B'].width = 25
            ws_bit.column_dimensions['C'].width = 70
            
            # Style data cells with alternating colors
            og_fill = xls.solid_fill(xls.COLOR_OG)
            oe_fill = xls.solid_fill(xls.COLOR_OE)
            
            for row in ws_bit.iter_rows(min_row=2):
                field = str(row[1].value).strip().upper() if row[1].value else ""
                if field == "OG":
                    for cell in row:
                        cell.fill = og_fill
                        cell.font = Font(bold=True, size=11)
                elif field.startswith("OE"):
                    for cell in row:
                        cell.fill = oe_fill
                
                row[1].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                row[2].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # 1. Format Planificación and Administración
        for sheet_name in ["Planificación", "Administración"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            
            # Freeze panes for UX (freeze row 1 and columns A, B)
            ws.freeze_panes = "C2"
            
            # Ensure headers
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = h
                
            # Column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 30
            ws.column_dimensions['I'].width = 30

            max_r = max(2, ws.max_row)

            # Native Tables for Premium look (Zebra stripes, auto filters)
            table_name = f"Tabla_{sheet_name.replace(' ', '')[:20]}"
            if table_name in ws.tables:
                ws.tables[table_name].ref = f"A1:I{max_r}"
            else:
                tab = Table(displayName=table_name, ref=f"A1:I{max_r}")
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                ws.add_table(tab)

            # Formato condicional: DataBar para el progreso y semáforo por Estado.
            estado_col = get_column_letter(IDX_ESTADO + 1)   # "F"
            prog_col = get_column_letter(IDX_PROGRESO + 1)    # "G"
            ws.conditional_formatting = type(ws.conditional_formatting)()
            ws.conditional_formatting.add(
                f"{prog_col}2:{prog_col}{max_r}",
                DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=xls.COLOR_DATABAR),
            )
            for status, color in xls.STATUS_COLORS.items():
                ws.conditional_formatting.add(
                    f"{estado_col}2:{estado_col}{max_r}",
                    CellIsRule(operator="equal", formula=[f'"{status}"'], fill=xls.solid_fill(color)),
                )

            # Validación de datos: dropdown de Estado (evita typos y normaliza el semáforo).
            ws.data_validations.dataValidation.clear()
            dv = DataValidation(
                type="list",
                formula1='"%s"' % ",".join(TASK_STATUSES),
                allow_blank=True,
            )
            dv.errorTitle = "Estado inválido"
            dv.error = "Selecciona un estado de la lista desplegable."
            dv.promptTitle = "Estado"
            dv.prompt = "Elige el estado de la tarea"
            ws.add_data_validation(dv)
            dv.add(f"{estado_col}2:{estado_col}{max_r}")

            # Cell formatting inside the table
            for row in ws.iter_rows(min_row=2, max_row=max_r, max_col=9):
                for cell in row:
                    if cell.column == IDX_PROGRESO + 1:
                        cell.number_format = '0%'
                    cell.alignment = align_left if cell.column in (2, 8, 9) else align_center
                    
        # 1.5. Format Evidencia
        if "Evidencia" in wb.sheetnames:
            ws_ev = wb["Evidencia"]
            
            # Freeze panes
            ws_ev.freeze_panes = "B2"
            
            ev_headers = ["Actividad", "Descripción", "Enlace / Ubicación"]
            for col_idx, h in enumerate(ev_headers, 1):
                cell = ws_ev.cell(row=1, column=col_idx)
                cell.value = h
            
            ws_ev.column_dimensions['A'].width = 12
            ws_ev.column_dimensions['B'].width = 40
            ws_ev.column_dimensions['C'].width = 60
            
            # Clean floating ghost rows
            last_valid_row = 1
            for row in ws_ev.iter_rows(min_row=2, max_col=3):
                if any(c.value for c in row):
                    last_valid_row = row[0].row
            if last_valid_row < ws_ev.max_row:
                ws_ev.delete_rows(last_valid_row + 1, ws_ev.max_row - last_valid_row)
                
            max_ev_row = max(2, ws_ev.max_row)
            
            # Native Table for Evidencia
            ev_table_name = "Tabla_Evidencia"
            if ev_table_name in ws_ev.tables:
                ws_ev.tables[ev_table_name].ref = f"A1:C{max_ev_row}"
            else:
                tab_ev = Table(displayName=ev_table_name, ref=f"A1:C{max_ev_row}")
                style_ev = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab_ev.tableStyleInfo = style_ev
                ws_ev.add_table(tab_ev)
            
            for row in ws_ev.iter_rows(min_row=2, max_row=max_ev_row, max_col=3):
                for cell in row:
                    cell.alignment = align_left if cell.column in (2, 3) else align_center
                link = row[2]
                if link.value and str(link.value).startswith(("http://", "https://")):
                    link.hyperlink = str(link.value)
                    link.font = Font(color="0563C1", underline="single")

        # 1.6 Dashboard con KPIs (vista de solo lectura).
        ValuelistExcelService._build_dashboard(wb)

        # 2. Draw Gantt Chart
        tasks = []
        min_date = None
        max_date = None
        
        for sheet_name in ["Planificación", "Administración"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                act_id = row[0].value
                desc = row[1].value
                start = row[3].value
                end = row[4].value
                if act_id and start and end:
                    # Parse dates
                    if isinstance(start, str):
                        try: start = datetime.strptime(start[:10], "%Y-%m-%d")
                        except: continue
                    if isinstance(end, str):
                        try: end = datetime.strptime(end[:10], "%Y-%m-%d")
                        except: continue
                        
                    if isinstance(start, datetime) and isinstance(end, datetime):
                        tasks.append({"id": act_id, "desc": desc, "start": start, "end": end})
                        if min_date is None or start < min_date: min_date = start
                        if max_date is None or end > max_date: max_date = end
        
        ws_gantt = wb["Gantt"] if "Gantt" in wb.sheetnames else wb["Carta Gantt"] if "Carta Gantt" in wb.sheetnames else None
        if not ws_gantt: return
        
        # Clear Gantt sheet
        ws_gantt.delete_rows(1, ws_gantt.max_row)
        ws_gantt.row_dimensions.clear()
        
        if not tasks or not min_date or not max_date:
            ws_gantt.cell(row=1, column=1).value = "No hay tareas con fechas válidas."
            return
            
        # Build timeline (weekly)
        current_date = min_date - timedelta(days=min_date.weekday())
        end_timeline = max_date + timedelta(days=(6 - max_date.weekday()))
        
        weeks = []
        while current_date <= end_timeline:
            weeks.append(current_date)
            current_date += timedelta(days=7)
            
        ws_gantt.cell(row=1, column=1).value = "ID"
        ws_gantt.cell(row=1, column=2).value = "Descripción"
        ws_gantt.cell(row=2, column=1).value = "ID"
        ws_gantt.cell(row=2, column=2).value = "Descripción"
        
        ws_gantt.column_dimensions['A'].width = 12
        ws_gantt.column_dimensions['B'].width = 30
        
        col_idx = 3
        current_month = None
        month_start_col = 3
        
        months_es = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
        
        for w in weeks:
            month_str = f"{months_es[w.month - 1]} {w.year}"
            if current_month != month_str:
                if current_month is not None:
                    if month_start_col < col_idx - 1:
                        ws_gantt.merge_cells(start_row=1, start_column=month_start_col, end_row=1, end_column=col_idx-1)
                    cell = ws_gantt.cell(row=1, column=month_start_col)
                    cell.value = current_month
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = thin_border
                current_month = month_str
                month_start_col = col_idx
                
            c2 = ws_gantt.cell(row=2, column=col_idx)
            c2.value = f"{w.day:02d}/{w.month:02d}"
            c2.fill = xls.solid_fill(xls.COLOR_SUBHEADER)
            c2.font = Font(bold=True)
            c2.alignment = align_center
            c2.border = thin_border
            ws_gantt.column_dimensions[get_column_letter(col_idx)].width = 8
            
            col_idx += 1
            
        if current_month is not None:
            if month_start_col < col_idx - 1:
                ws_gantt.merge_cells(start_row=1, start_column=month_start_col, end_row=1, end_column=col_idx-1)
            cell = ws_gantt.cell(row=1, column=month_start_col)
            cell.value = current_month
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border

        # Draw tasks
        gantt_fill = xls.solid_fill(xls.COLOR_GANTT_BAR)
        
        row_idx = 3
        for t in sorted(tasks, key=lambda x: x["start"]):
            c_id = ws_gantt.cell(row=row_idx, column=1)
            c_id.value = t["id"]
            c_id.alignment = align_center
            c_id.border = thin_border
            
            c_desc = ws_gantt.cell(row=row_idx, column=2)
            c_desc.value = t["desc"]
            c_desc.alignment = align_left
            c_desc.border = thin_border
            
            c_idx = 3
            for w in weeks:
                w_end = w + timedelta(days=6)
                c_cell = ws_gantt.cell(row=row_idx, column=c_idx)
                if t["start"] <= w_end and t["end"] >= w:
                    c_cell.fill = gantt_fill
                c_cell.border = thin_border
                c_idx += 1
                
            row_idx += 1

    @staticmethod
    def _build_dashboard(wb: openpyxl.Workbook) -> None:
        """(Re)genera la hoja Dashboard con KPIs de solo lectura.

        Une el estado de las tareas de planificación en una vista ejecutiva:
        avance global, conteo por estado y tareas próximas a vencer.
        """
        from datetime import datetime, date, timedelta
        from openpyxl.styles import Font
        from src.infrastructure import excel_styles as xls

        # 1. Recolectar tareas de las hojas de planificación.
        tasks = []
        for sheet_name in ["Planificación", "Administración"]:
            if sheet_name not in wb.sheetnames:
                continue
            for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
                if len(row) <= IDX_PROGRESO or not row[IDX_ACTIVIDAD]:
                    continue
                progress = to_fraction(row[IDX_PROGRESO])
                estado_raw = str(row[IDX_ESTADO]).strip().upper() if row[IDX_ESTADO] else ""
                estado = estado_raw if estado_raw in TASK_STATUSES else derive_status(progress)
                tasks.append({
                    "id": str(row[IDX_ACTIVIDAD]),
                    "estado": estado,
                    "progress": progress,
                    "fin": row[IDX_FIN],
                })

        total = len(tasks)
        counts = {s: 0 for s in TASK_STATUSES}
        for t in tasks:
            counts[t["estado"]] = counts.get(t["estado"], 0) + 1
        avg = (sum(t["progress"] for t in tasks) / total) if total else 0.0

        # 2. Próximas a vencer (7 días) y no completadas.
        def _as_date(value):
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            try:
                return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
            except (ValueError, TypeError):
                return None

        horizon = date.today() + timedelta(days=7)
        upcoming = [
            (t["id"], d)
            for t in tasks
            if (d := _as_date(t["fin"])) is not None and d <= horizon and t["progress"] < 1.0
        ]
        upcoming.sort(key=lambda x: x[1])

        # 3. (Re)crear la hoja, limpiando merges previos para evitar corrupción.
        if "Dashboard" in wb.sheetnames:
            ws = wb["Dashboard"]
            for merged in list(ws.merged_cells.ranges):
                ws.unmerge_cells(str(merged))
            if ws.max_row:
                ws.delete_rows(1, ws.max_row)
            ws.row_dimensions.clear()
        else:
            ws = wb.create_sheet("Dashboard", 0)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 16
        ws.sheet_view.showGridLines = False

        ws["A1"] = "📊 Dashboard del Proyecto"
        ws.merge_cells("A1:B1")
        title = ws["A1"]
        title.fill = xls.header_fill()
        title.font = xls.header_font()
        title.alignment = xls.align_center()

        kpis = [
            ("Avance global", avg, "0%"),
            ("Total de tareas", total, "0"),
            ("✅ Completadas", counts.get(STATUS_DONE, 0), "0"),
            ("🟡 En curso", counts.get(STATUS_IN_PROGRESS, 0), "0"),
            ("⬜ No comenzadas", counts.get(STATUS_NOT_STARTED, 0), "0"),
            ("🔴 Bloqueadas", counts.get(STATUS_BLOCKED, 0), "0"),
        ]
        r = 3
        for label, value, fmt in kpis:
            c_label = ws.cell(row=r, column=1, value=label)
            c_label.font = Font(bold=True)
            c_label.border = xls.thin_border()
            c_value = ws.cell(row=r, column=2, value=value)
            c_value.number_format = fmt
            c_value.alignment = xls.align_center()
            c_value.border = xls.thin_border()
            r += 1

        r += 1
        head = ws.cell(row=r, column=1, value="⏰ Próximas a vencer (7 días)")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        head.fill = xls.header_fill()
        head.font = xls.header_font()
        head.alignment = xls.align_center()
        r += 1
        ws.cell(row=r, column=1, value="Tarea").font = Font(bold=True)
        ws.cell(row=r, column=2, value="Vence").font = Font(bold=True)
        r += 1
        if upcoming:
            for tid, due in upcoming:
                ws.cell(row=r, column=1, value=tid)
                ws.cell(row=r, column=2, value=due.strftime("%d/%m/%Y")).alignment = xls.align_center()
                r += 1
        else:
            ws.cell(row=r, column=1, value="Sin tareas próximas a vencer 🎉")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    async def get_my_tasks(self, target_name: str) -> list[dict[str, Any]]:
        """Busca las tareas asignadas al usuario en Planificación y Administración."""
        if not target_name:
            return []

        def _read() -> list[dict[str, Any]]:
            try:
                wb = openpyxl.load_workbook(self._excel_path, data_only=True)
                tasks = []
                for sheet_name in ["Planificación", "Administración"]:
                    if sheet_name not in wb.sheetnames:
                        continue
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        # Cols: 0=Act, 1=Desc, 2=Resp, 3=Inicio, 4=Fin, 5=Estado, 6=%logro
                        act_id = row[IDX_ACTIVIDAD]
                        resp = row[IDX_RESP]

                        if act_id and (str(act_id).startswith("A") or str(act_id).startswith("AD")) and resp == target_name:
                            tasks.append({
                                "id": str(act_id),
                                "desc": str(row[IDX_DESC]) if row[IDX_DESC] else "",
                                "progress": to_fraction(row[IDX_PROGRESO]),
                                "estado": str(row[IDX_ESTADO]).strip() if row[IDX_ESTADO] else "",
                            })
                return tasks
            except Exception:
                return []

        return await asyncio.to_thread(_read)

    async def get_bitacora_summary(self) -> dict[str, Any]:
        """Lee el OG y los OE de la Hoja 1 (Bitácora)."""
        def _read() -> dict[str, Any]:
            try:
                wb = openpyxl.load_workbook(self._excel_path, data_only=True)
                ws = wb["Bitácora"]
                
                summary = {"proyecto": "", "og": "", "oe": []}
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    col_b = row[1]
                    col_c = row[2]
                    if not col_b:
                        continue
                        
                    b_str = str(col_b).strip().upper()
                    c_str = str(col_c).strip() if col_c else ""
                    
                    if b_str == "PROYECTO":
                        summary["proyecto"] = c_str
                    elif b_str == "OG":
                        summary["og"] = c_str
                    elif b_str.startswith("OE"):
                        # Keep even if empty so it can be edited/deleted
                        summary["oe"].append({"id": b_str, "desc": c_str})
                            
                return summary
            except Exception:
                return {"proyecto": "", "og": "", "oe": []}

        return await asyncio.to_thread(_read)

    async def update_task_progress(self, task_id: str, progress: float) -> bool:
        """Actualiza el % de logro de una tarea y autoderiva su Estado.

        El progreso se normaliza a fracción 0-1 y el Estado se deriva del
        progreso (respetando BLOQUEADO manual salvo que llegue al 100%).
        """
        fraction = to_fraction(progress)

        def _write() -> bool:
            try:
                wb = openpyxl.load_workbook(self._excel_path)
                for sheet_name in ["Planificación", "Administración"]:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2):
                        if row[IDX_ACTIVIDAD].value and str(row[IDX_ACTIVIDAD].value) == task_id:
                            row[IDX_PROGRESO].value = fraction
                            current_estado = row[IDX_ESTADO].value
                            row[IDX_ESTADO].value = derive_status(fraction, current_estado)
                            self._apply_gantt_and_styles(wb)
                            wb.save(self._excel_path)
                            return True
                return False
            except Exception:
                return False

        return await asyncio.to_thread(_write)

    async def add_evidence(self, task_id: str, url: str) -> bool:
        """Añade un enlace de evidencia a la Hoja 5."""
        def _write() -> bool:
            try:
                wb = openpyxl.load_workbook(self._excel_path)
                
                # Primero buscar la descripción en Planificación
                ws_plan = wb["Planificación"]
                desc = ""
                for row in ws_plan.iter_rows(min_row=2):
                    cell_id = row[0]
                    if cell_id.value and str(cell_id.value) == task_id:
                        desc = str(row[1].value) if row[1].value else ""
                        break
                        
                if not desc:
                    return False
                    
                ws_evi = wb["Evidencia"]
                
                # Append manually to the first truly empty row to avoid ghost rows
                last_row = 1
                for i in range(2, ws_evi.max_row + 2):
                    if not ws_evi.cell(row=i, column=1).value and not ws_evi.cell(row=i, column=2).value:
                        last_row = i
                        break
                ws_evi.cell(row=last_row, column=1).value = task_id
                ws_evi.cell(row=last_row, column=2).value = desc
                ws_evi.cell(row=last_row, column=3).value = url
                
                self._apply_gantt_and_styles(wb)
                wb.save(self._excel_path)
                return True
            except Exception:
                return False

        return await asyncio.to_thread(_write)

    async def create_task(self, oe_id: str, desc: str, resp: str, start: str, end: str, entregable: str = "", comentarios: str = "") -> bool:
        """Agrega una nueva fila al final de la hoja correspondiente, autogenerando el ID."""
        def _write() -> bool:
            try:
                wb = openpyxl.load_workbook(self._excel_path)
                sheet_name = "Administración" if oe_id == "AD" else "Planificación"
                ws = wb[sheet_name]
                
                # Determine prefix
                if oe_id == "AD":
                    prefix = "AD"
                elif oe_id == "A":
                    prefix = "A0."
                elif oe_id.startswith("OE"):
                    num = oe_id.replace("OE", "").strip()
                    prefix = f"A{num}."
                else:
                    prefix = "A"

                # Find max suffix
                max_suffix = 0
                for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    cell_val = r[0].value
                    if cell_val and str(cell_val).startswith(prefix):
                        suffix_str = str(cell_val)[len(prefix):]
                        try:
                            max_suffix = max(max_suffix, int(suffix_str))
                        except ValueError:
                            pass
                
                act_id = f"{prefix}{max_suffix + 1}"
                
                # Encontrar la primera fila libre en la columna A
                last_row = 1
                for i in range(2, ws.max_row + 2):
                    if not ws.cell(row=i, column=1).value:
                        last_row = i
                        break
                
                from datetime import datetime
                try:
                    start_dt = datetime.strptime(start, "%Y-%m-%d")
                    end_dt = datetime.strptime(end, "%Y-%m-%d")
                except ValueError:
                    start_dt = start
                    end_dt = end

                ws.cell(row=last_row, column=1).value = act_id
                ws.cell(row=last_row, column=2).value = desc
                ws.cell(row=last_row, column=3).value = resp
                ws.cell(row=last_row, column=4).value = start_dt
                ws.cell(row=last_row, column=5).value = end_dt
                ws.cell(row=last_row, column=IDX_ESTADO + 1).value = STATUS_NOT_STARTED
                ws.cell(row=last_row, column=IDX_PROGRESO + 1).value = 0.0
                ws.cell(row=last_row, column=8).value = entregable
                ws.cell(row=last_row, column=9).value = comentarios
                self._apply_gantt_and_styles(wb)
                wb.save(self._excel_path)
                return True
            except Exception as e:
                print(f"Error create_task: {e}")
                return False

        return await asyncio.to_thread(_write)

    async def generate_gantt(self) -> str:
        """Genera un diagrama Mermaid a partir de las fechas en la Hoja 3."""
        def _read() -> str:
            try:
                wb = openpyxl.load_workbook(self._excel_path, data_only=True)
                ws = wb["Planificación"]
                
                lines = [
                    "```mermaid",
                    "gantt",
                    "    title Cronograma de Planificación (Excel)",
                    "    dateFormat  YYYY-MM-DD",
                    "    axisFormat  %d-%b"
                ]
                
                for row in ws.iter_rows(min_row=2):
                    act_id = row[0].value
                    desc = row[1].value
                    start = row[3].value
                    end = row[4].value
                    
                    if act_id and str(act_id).startswith("A") and start and end:
                        # Clean dates if they are datetime objects
                        from datetime import datetime
                        if isinstance(start, datetime): start_str = start.strftime("%Y-%m-%d")
                        else: start_str = str(start)[:10]
                        
                        if isinstance(end, datetime): end_str = end.strftime("%Y-%m-%d")
                        else: end_str = str(end)[:10]
                        
                        safe_desc = str(desc).replace(":", "") if desc else "Tarea"
                        lines.append(f"    {act_id} {safe_desc[:30]} : {act_id}, {start_str}, {end_str}")
                        
                if len(lines) == 5:
                    return "⚠️ No hay tareas con fechas de inicio y fin en la Planificación para graficar."
                    
                lines.append("```")
                return "\n".join(lines)
            except Exception as e:
                return f"Error generando Gantt: {e}"

        return await asyncio.to_thread(_read)

    async def get_task_by_id(self, task_id: str) -> dict[str, Any] | None:
        """Busca una tarea en Hoja 3 o Hoja 4 y devuelve sus detalles actuales."""
        def _read() -> dict[str, Any] | None:
            try:
                wb = openpyxl.load_workbook(self._excel_path, data_only=True)
                
                for sheet_name in ["Planificación", "Administración"]:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        act_id = row[0]
                        if act_id and str(act_id) == task_id:
                            from datetime import datetime
                            def _format_date(d):
                                if isinstance(d, datetime): return d.strftime("%Y-%m-%d")
                                return str(d)[:10] if d else ""
                                
                            return {
                                "id": str(act_id),
                                "desc": str(row[1]) if row[1] else "",
                                "resp": str(row[2]) if row[2] else "",
                                "start": _format_date(row[3]),
                                "end": _format_date(row[4]),
                                "entregable": str(row[7]) if len(row) > 7 and row[7] else "",
                                "comentarios": str(row[8]) if len(row) > 8 and row[8] else ""
                            }
                return None
            except Exception:
                return None

        return await asyncio.to_thread(_read)

    async def update_task_details(self, task_id: str, desc: str, resp: str, start: str, end: str, entregable: str = "", comentarios: str = "") -> bool:
        """Sobrescribe los detalles de una tarea existente en Hoja 3 o 4."""
        def _write() -> bool:
            try:
                wb = openpyxl.load_workbook(self._excel_path)
                
                from datetime import datetime
                try:
                    start_dt = datetime.strptime(start, "%Y-%m-%d")
                    end_dt = datetime.strptime(end, "%Y-%m-%d")
                except ValueError:
                    start_dt = start
                    end_dt = end
                
                for sheet_name in ["Planificación", "Administración"]:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2):
                        cell_id = row[0]
                        if cell_id.value and str(cell_id.value) == task_id:
                            row[1].value = desc
                            row[2].value = resp
                            row[3].value = start_dt
                            row[4].value = end_dt
                            # Ensure row has enough cells
                            while len(row) < 9:
                                ws.cell(row=cell_id.row, column=len(row)+1).value = ""
                                row = tuple(ws.iter_rows(min_row=cell_id.row, max_row=cell_id.row))[0]
                            row[7].value = entregable
                            row[8].value = comentarios
                            self._apply_gantt_and_styles(wb)
                            wb.save(self._excel_path)
                            return True
                return False
            except Exception:
                return False

        return await asyncio.to_thread(_write)

    async def delete_task_by_id(self, task_id: str) -> bool:
        """Elimina una fila de la Hoja 3 o 4."""
        def _write() -> bool:
            try:
                wb = openpyxl.load_workbook(self._excel_path)
                
                for sheet_name in ["Planificación", "Administración"]:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2):
                        cell_id = row[0]
                        if cell_id.value and str(cell_id.value) == task_id:
                            ws.delete_rows(cell_id.row)
                            self._apply_gantt_and_styles(wb)
                            wb.save(self._excel_path)
                            return True
                return False
            except Exception:
                return False

        return await asyncio.to_thread(_write)

    async def update_bitacora_full(self, updates: dict[str, str], new_oe: str = "") -> bool:
        """Actualiza todos los objetivos de la Bitácora dinámicamente, agrega un nuevo OE si se provee, y aplica estilos."""
        def _write() -> bool:
            try:
                wb = openpyxl.load_workbook(self._excel_path)
                ws = wb["Bitácora"]
                
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=12)
                bold_font = Font(bold=True)
                align_center = Alignment(horizontal="center", vertical="center")
                align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))

                # 1. Extract current data in case something is missing in updates
                current_data = {"PROYECTO": "", "OG": ""}
                oes = []
                for r in range(2, ws.max_row + 1):
                    b = ws.cell(row=r, column=2).value
                    c = ws.cell(row=r, column=3).value
                    if b:
                        b_str = str(b).strip().upper()
                        c_str = str(c).strip() if c else ""
                        if b_str in ["PROYECTO", "OG"]:
                            current_data[b_str] = c_str
                        elif b_str.startswith("OE"):
                            oes.append({"id": b_str, "desc": c_str})
                
                # 2. Apply updates
                if "PROYECTO" in updates: current_data["PROYECTO"] = updates["PROYECTO"]
                if "OG" in updates: current_data["OG"] = updates["OG"]
                
                new_oes = []
                for oe in oes:
                    oe_id = oe["id"]
                    if oe_id in updates:
                        desc = str(updates[oe_id]).strip()
                        if desc: # If not empty, we keep it
                            new_oes.append({"id": oe_id, "desc": desc})
                    else:
                        if oe["desc"]: # Keep if it had text
                            new_oes.append(oe)
                            
                # 3. Add new OE if provided
                if new_oe.strip():
                    last_num = 0
                    for oe in new_oes:
                        try:
                            num = int(oe["id"].replace("OE", "").strip())
                            last_num = max(last_num, num)
                        except ValueError:
                            pass
                    new_oes.append({"id": f"OE{last_num + 1}", "desc": new_oe.strip()})
                    
                # 4. Clear existing rows and rewrite everything sequentially
                ws.delete_rows(2, ws.max_row)
                ws.row_dimensions.clear()
                
                ws.append(["", "PROYECTO", current_data["PROYECTO"]])
                ws.append(["", "OG", current_data["OG"]])
                for oe in new_oes:
                    ws.append(["", oe["id"], oe["desc"]])

                # 5. Apply Premium Styles
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 80
                
                # Header Style
                ws.cell(row=1, column=2).value = "ID Objetivo"
                ws.cell(row=1, column=3).value = "Descripción"
                for col in (2, 3):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = thin_border
                
                # Content Style
                for row_idx in range(2, ws.max_row + 1):
                    val_b = ws.cell(row=row_idx, column=2).value
                    if val_b:
                        c_b = ws.cell(row=row_idx, column=2)
                        c_b.font = bold_font
                        c_b.alignment = align_center
                        c_b.border = thin_border
                        
                        c_c = ws.cell(row=row_idx, column=3)
                        c_c.alignment = align_left
                        c_c.border = thin_border
                
                wb.save(self._excel_path)
                return True
            except Exception as e:
                print(f"Error updating bitacora: {e}")
                return False

        return await asyncio.to_thread(_write)

    async def get_objective_progress(self) -> list[dict[str, Any]]:
        """Avance agregado por Objetivo Específico (y Administración).

        Agrupa las tareas por su prefijo de OE (A1.x → OE1, AD.x →
        Administración) y calcula avance promedio, total y completadas,
        tomando la descripción del OE desde la Bitácora.
        """
        def _read() -> list[dict[str, Any]]:
            from collections import defaultdict
            try:
                wb = openpyxl.load_workbook(self._excel_path, data_only=True)
            except Exception:
                return []

            oe_desc: dict[str, str] = {}
            if "Bitácora" in wb.sheetnames:
                for row in wb["Bitácora"].iter_rows(min_row=2, values_only=True):
                    if len(row) < 3 or not row[1]:
                        continue
                    key = str(row[1]).strip().upper()
                    if key.startswith("OE"):
                        oe_desc[key] = str(row[2]).strip() if row[2] else ""

            groups: dict[str, dict[str, float]] = defaultdict(
                lambda: {"total": 0, "done": 0, "sum": 0.0}
            )

            def _objective_key(act_id: Any) -> str:
                s = str(act_id).strip()
                if s.upper().startswith("AD"):
                    return "Administración"
                match = re.match(r"A(\d+)\.", s)
                return f"OE{match.group(1)}" if match else "Otros"

            for sheet_name in ["Planificación", "Administración"]:
                if sheet_name not in wb.sheetnames:
                    continue
                for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
                    if len(row) <= IDX_PROGRESO or not row[IDX_ACTIVIDAD]:
                        continue
                    key = _objective_key(row[IDX_ACTIVIDAD])
                    progress = to_fraction(row[IDX_PROGRESO])
                    bucket = groups[key]
                    bucket["total"] += 1
                    bucket["sum"] += progress
                    if progress >= 1.0:
                        bucket["done"] += 1

            def _sort_key(k: str):
                match = re.match(r"OE(\d+)", k)
                return (0, int(match.group(1))) if match else (1, k)

            result = []
            for key in sorted(groups.keys(), key=_sort_key):
                bucket = groups[key]
                total = int(bucket["total"])
                avg = bucket["sum"] / total if total else 0.0
                result.append({
                    "id": key,
                    "desc": oe_desc.get(key, ""),
                    "progress": avg,
                    "total": total,
                    "done": int(bucket["done"]),
                })
            return result

        return await asyncio.to_thread(_read)

    async def get_all_active_tasks(self) -> dict[str, list[str]]:
        """Devuelve las tareas no completadas agrupadas por responsable."""
        def _read() -> dict[str, list[str]]:
            from collections import defaultdict
            grouped = defaultdict(list)
            try:
                wb = openpyxl.load_workbook(self._excel_path, data_only=True)

                for sheet_name in ["Planificación", "Administración"]:
                    if sheet_name not in wb.sheetnames:
                        continue
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if len(row) <= IDX_PROGRESO or not row[IDX_ACTIVIDAD]:
                            continue

                        act_id = row[IDX_ACTIVIDAD]
                        resp = str(row[IDX_RESP]).strip() if row[IDX_RESP] else "Sin asignar"
                        progress = to_fraction(row[IDX_PROGRESO])

                        # El progreso se guarda como fracción 0-1; 1.0 = completada.
                        if progress >= 1.0:
                            continue

                        desc = row[IDX_DESC] or ""
                        grouped[resp].append(f"• *{act_id}*: {desc} ({progress * 100:.0f}%)")

                return dict(grouped)
            except Exception:
                return {}

        return await asyncio.to_thread(_read)

    async def get_all_edit_options(self) -> dict[str, list[dict[str, str]]]:
        """Devuelve opciones agrupadas de Tareas para el selector de Slack."""
        def _read() -> dict[str, list[dict[str, str]]]:
            options = {"tareas": [], "bitacora": []}
            try:
                wb = openpyxl.load_workbook(self._excel_path, data_only=True)
                
                # Tareas
                for sheet_name in ["Planificación", "Administración"]:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        act_id = row[0]
                        if act_id:
                            desc = str(row[1])[:50] if row[1] else "Sin descripción"
                            options["tareas"].append({
                                "text": {"type": "plain_text", "text": f"{act_id} - {desc}"},
                                "value": f"tarea|{act_id}"
                            })
                    
                return options
            except Exception:
                return options

        return await asyncio.to_thread(_read)
