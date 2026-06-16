import openpyxl

wb = openpyxl.load_workbook("project_tracking.xlsx", data_only=False)
if "Gantt" in wb.sheetnames:
    ws = wb["Gantt"]
    for i in range(3, 10):
        fills = [ws.cell(row=i, column=c).fill.start_color.index for c in range(1, 10) if ws.cell(row=i, column=c).fill]
        print(f"Row {i} fills: {fills}")
