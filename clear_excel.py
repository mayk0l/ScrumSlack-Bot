import openpyxl

wb = openpyxl.load_workbook("project_tracking.xlsx")

# Clear Planificación
ws = wb["Planificación"]
ws.delete_rows(2, ws.max_row)

# Clear Administración
ws = wb["Administración"]
ws.delete_rows(2, ws.max_row)

# Clear Evidencias
if "Evidencias" in wb.sheetnames:
    ws = wb["Evidencias"]
    ws.delete_rows(2, ws.max_row)

# Clear Bitácora
ws = wb["Bitácora"]
for row in ws.iter_rows(min_row=2):
    row[1].value = None

wb.save("project_tracking.xlsx")
print("Cleaned Excel.")
