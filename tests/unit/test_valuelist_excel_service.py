"""Módulo: test_valuelist_excel_service.

Tests del servicio de Excel usando archivos .xlsx reales temporales en lugar de
mocks, para que sean robustos frente a refactors internos de openpyxl.
"""

import openpyxl
import pytest
from datetime import datetime

from src.application.valuelist_excel_service import (
    ValuelistExcelService,
    STATUS_NOT_STARTED,
    STATUS_IN_PROGRESS,
    STATUS_DONE,
    STATUS_BLOCKED,
    to_fraction,
    derive_status,
)

# Esquema canónico de columnas de las hojas de tareas (Planificación / Administración).
HEADERS_TASKS = [
    "Actividad",
    "Descripción (resumen)",
    "Responsable",
    "Comienzo",
    "Fin (Esperado/logrado)",
    "Estado",
    "% de logro",
    "Entregable",
    "Comentarios",
]


def _build_workbook(path: str) -> None:
    """Crea un libro de prueba con la estructura real del tracking."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_b = wb.create_sheet("Bitácora")
    ws_b.append(["", "ID Objetivo", "Descripción"])
    ws_b.append(["", "PROYECTO", "Proyecto X"])
    ws_b.append(["", "OG", "Objetivo general"])
    ws_b.append(["", "OE1", "Primer objetivo"])

    # La hoja Gantt es necesaria porque _apply_gantt_and_styles la usa al escribir.
    wb.create_sheet("Gantt")

    ws_p = wb.create_sheet("Planificación")
    ws_p.append(HEADERS_TASKS)
    ws_p.append(["A1.1", "Backend", "Ana", datetime(2026, 6, 1), datetime(2026, 6, 5), "EN CURSO", 0.5, "", ""])
    ws_p.append(["A1.2", "Frontend", "Beto", datetime(2026, 6, 2), datetime(2026, 6, 8), "NO COMENZADO", 0.0, "", ""])
    ws_p.append(["A1.3", "Base de datos", "Ana", datetime(2026, 6, 3), datetime(2026, 6, 9), "COMPLETADO", 1.0, "", ""])

    ws_a = wb.create_sheet("Administración")
    ws_a.append(HEADERS_TASKS)

    ws_e = wb.create_sheet("Evidencia")
    ws_e.append(["Actividad", "Descripción", "Enlace / Ubicación"])

    wb.save(path)


@pytest.fixture
def excel_path(tmp_path) -> str:
    p = tmp_path / "tracking.xlsx"
    _build_workbook(str(p))
    return str(p)


@pytest.fixture
def service(excel_path) -> ValuelistExcelService:
    return ValuelistExcelService(excel_path=excel_path)


@pytest.mark.asyncio
async def test_get_my_tasks(service):
    tasks = await service.get_my_tasks("Ana")
    ids = {t["id"] for t in tasks}
    assert ids == {"A1.1", "A1.3"}
    a11 = next(t for t in tasks if t["id"] == "A1.1")
    assert a11["desc"] == "Backend"
    assert a11["progress"] == 0.5


@pytest.mark.asyncio
async def test_get_my_tasks_unknown_user(service):
    assert await service.get_my_tasks("Nadie") == []


@pytest.mark.asyncio
async def test_get_bitacora_summary(service):
    summary = await service.get_bitacora_summary()
    assert summary["proyecto"] == "Proyecto X"
    assert summary["og"] == "Objetivo general"
    assert len(summary["oe"]) == 1
    assert summary["oe"][0]["id"] == "OE1"
    assert summary["oe"][0]["desc"] == "Primer objetivo"


@pytest.mark.asyncio
async def test_update_task_progress(service, excel_path):
    assert await service.update_task_progress("A1.2", 1.0) is True
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Planificación"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == "A1.2":
            assert row[6] == 1.0
            break
    else:
        pytest.fail("No se encontró A1.2")


@pytest.mark.asyncio
async def test_update_task_progress_not_found(service):
    assert await service.update_task_progress("ZZZ", 0.5) is False


def _read_task_row(excel_path, sheet, task_id):
    wb = openpyxl.load_workbook(excel_path)
    for row in wb[sheet].iter_rows(min_row=2, values_only=True):
        if row[0] == task_id:
            return row
    return None


@pytest.mark.asyncio
async def test_update_progress_derives_status(service, excel_path):
    await service.update_task_progress("A1.2", 0.5)
    row = _read_task_row(excel_path, "Planificación", "A1.2")
    assert row[5] == STATUS_IN_PROGRESS
    assert row[6] == 0.5

    await service.update_task_progress("A1.2", 1.0)
    assert _read_task_row(excel_path, "Planificación", "A1.2")[5] == STATUS_DONE

    await service.update_task_progress("A1.2", 0.0)
    assert _read_task_row(excel_path, "Planificación", "A1.2")[5] == STATUS_NOT_STARTED


@pytest.mark.asyncio
async def test_update_progress_normalizes_percentage(service, excel_path):
    # 50 (porcentaje) debe almacenarse como 0.5 (fracción).
    await service.update_task_progress("A1.1", 50)
    assert _read_task_row(excel_path, "Planificación", "A1.1")[6] == 0.5


@pytest.mark.asyncio
async def test_update_progress_preserves_blocked(service, excel_path):
    wb = openpyxl.load_workbook(excel_path)
    for r in wb["Planificación"].iter_rows(min_row=2):
        if r[0].value == "A1.1":
            r[5].value = STATUS_BLOCKED
    wb.save(excel_path)

    # Un avance parcial no debe quitar el bloqueo...
    await service.update_task_progress("A1.1", 0.3)
    assert _read_task_row(excel_path, "Planificación", "A1.1")[5] == STATUS_BLOCKED

    # ...pero llegar al 100% sí lo completa.
    await service.update_task_progress("A1.1", 1.0)
    assert _read_task_row(excel_path, "Planificación", "A1.1")[5] == STATUS_DONE


def test_to_fraction_and_derive_status():
    assert to_fraction(50) == 0.5
    assert to_fraction(0.5) == 0.5
    assert to_fraction("75%") == 0.75
    assert to_fraction(None) == 0.0
    assert to_fraction(150) == 1.0
    assert derive_status(0.0) == STATUS_NOT_STARTED
    assert derive_status(0.4) == STATUS_IN_PROGRESS
    assert derive_status(1.0) == STATUS_DONE
    assert derive_status(0.4, STATUS_BLOCKED) == STATUS_BLOCKED
    assert derive_status(1.0, STATUS_BLOCKED) == STATUS_DONE


@pytest.mark.asyncio
async def test_create_task(service, excel_path):
    ok = await service.create_task("OE1", "Nueva tarea", "Ana", "2026-06-10", "2026-06-12")
    assert ok is True
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Planificación"]
    descs = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "Nueva tarea" in descs


@pytest.mark.asyncio
async def test_add_evidence(service, excel_path):
    ok = await service.add_evidence("A1.1", "http://github.com/pr/1")
    assert ok is True
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Evidencia"]
    rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    assert any(r[0] == "A1.1" and r[2] == "http://github.com/pr/1" for r in rows)


@pytest.mark.asyncio
async def test_add_evidence_unknown_task(service):
    assert await service.add_evidence("ZZZ", "http://x") is False


@pytest.mark.asyncio
async def test_get_task_by_id(service):
    task = await service.get_task_by_id("A1.1")
    assert task is not None
    assert task["id"] == "A1.1"
    assert task["desc"] == "Backend"
    assert task["resp"] == "Ana"


@pytest.mark.asyncio
async def test_get_task_by_id_not_found(service):
    assert await service.get_task_by_id("ZZZ") is None


@pytest.mark.asyncio
async def test_update_task_details(service, excel_path):
    ok = await service.update_task_details("A1.1", "Backend v2", "Beto", "2026-06-10", "2026-06-15")
    assert ok is True
    task = await service.get_task_by_id("A1.1")
    assert task["desc"] == "Backend v2"
    assert task["resp"] == "Beto"


@pytest.mark.asyncio
async def test_delete_task_by_id(service, excel_path):
    assert await service.delete_task_by_id("A1.2") is True
    wb = openpyxl.load_workbook(excel_path)
    ids = [row[0] for row in wb["Planificación"].iter_rows(min_row=2, values_only=True)]
    assert "A1.2" not in ids


@pytest.mark.asyncio
async def test_update_bitacora_full(service):
    ok = await service.update_bitacora_full({"OG": "Nuevo OG"}, new_oe="Segundo objetivo")
    assert ok is True
    summary = await service.get_bitacora_summary()
    assert summary["og"] == "Nuevo OG"
    oe_ids = {oe["id"] for oe in summary["oe"]}
    assert "OE1" in oe_ids
    assert "OE2" in oe_ids


@pytest.mark.asyncio
async def test_get_all_active_tasks_groups_by_responsible(service):
    grouped = await service.get_all_active_tasks()
    assert "Ana" in grouped
    assert "Beto" in grouped
    assert any("A1.1" in line for line in grouped["Ana"])


@pytest.mark.asyncio
async def test_generate_gantt(service):
    gantt = await service.generate_gantt()
    assert "gantt" in gantt
    assert "A1.1" in gantt


@pytest.mark.asyncio
async def test_estado_header_is_canonical(service, excel_path):
    # Un write dispara _apply_gantt_and_styles, que reescribe los headers.
    await service.create_task("OE1", "Otra", "Ana", "2026-06-10", "2026-06-12")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Planificación"]
    headers = [c.value for c in ws[1]]
    assert headers[5] == "Estado"
    assert headers[6] == "% de logro"


@pytest.mark.asyncio
async def test_estado_dropdown_and_conditional_formatting(service, excel_path):
    await service.create_task("OE1", "Con estado", "Ana", "2026-06-10", "2026-06-12")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Planificación"]
    # Dropdown de Estado con los 4 valores válidos.
    formulas = [dv.formula1 or "" for dv in ws.data_validations.dataValidation]
    assert any("EN CURSO" in f and "BLOQUEADO" in f for f in formulas)
    # Semáforo: hay reglas de formato condicional aplicadas.
    assert len(list(ws.conditional_formatting)) >= 1


@pytest.mark.asyncio
async def test_evidence_is_hyperlink(service, excel_path):
    await service.add_evidence("A1.1", "https://github.com/org/repo/pull/1")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Evidencia"]
    found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "A1.1":
            assert row[2].hyperlink is not None
            found = True
    assert found


@pytest.mark.asyncio
async def test_dashboard_sheet_with_kpis(service, excel_path):
    await service.update_task_progress("A1.1", 1.0)
    wb = openpyxl.load_workbook(excel_path)
    assert "Dashboard" in wb.sheetnames
    texts = []
    for row in wb["Dashboard"].iter_rows(values_only=True):
        texts.extend(str(v) for v in row if v is not None)
    joined = " ".join(texts)
    assert "Avance global" in joined
    assert "Total de tareas" in joined
    assert "Próximas a vencer" in joined

