import asyncio
import re
import openpyxl
from slack_sdk.web.async_client import AsyncWebClient
from src.config import settings

async def main():
    print("Iniciando corrección de responsables en Excel...")
    try:
        wb = openpyxl.load_workbook("project_tracking.xlsx")
        client = AsyncWebClient(token=settings.slack_bot_token)
        
        user_cache = {}
        
        changed = False
        for sheet_name in ["Planificación", "Administración"]:
            if sheet_name not in wb.sheetnames:
                continue
                
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2):
                val = row[2].value
                if val and isinstance(val, str) and val.startswith("<@") and val.endswith(">"):
                    user_id = val[2:-1]
                    if user_id not in user_cache:
                        try:
                            user_info = await client.users_info(user=user_id)
                            real_name = user_info["user"].get("real_name") or user_info["user"].get("name")
                            user_cache[user_id] = real_name
                        except Exception as e:
                            print(f"Error fetching user {user_id}: {e}")
                            user_cache[user_id] = val # fallback
                            
                    real_name = user_cache[user_id]
                    if real_name != val:
                        row[2].value = real_name
                        changed = True
                        print(f"[{sheet_name}] Corregido {val} -> {real_name}")
                        
        if changed:
            wb.save("project_tracking.xlsx")
            print("Excel guardado con éxito.")
        else:
            print("No se encontraron IDs para corregir.")
            
    except Exception as e:
        print(f"Error general: {e}")

if __name__ == "__main__":
    asyncio.run(main())
