import openpyxl
import shutil

# Start fresh
shutil.copy("excel/Bitacora-Rentabilidad-Valuelist.xlsx", "project_tracking.xlsx")

wb = openpyxl.load_workbook("project_tracking.xlsx")

# 1. Clean Planificación
ws = wb["Planificación"]
max_r = ws.max_row
if max_r >= 2:
    ws.delete_rows(2, max_r)

# 2. Clean Administración
ws = wb["Administración"]
max_r = ws.max_row
if max_r >= 2:
    ws.delete_rows(2, max_r)

# 3. Clean Evidencias
if "Evidencias" in wb.sheetnames:
    ws = wb["Evidencias"]
    max_r = ws.max_row
    if max_r >= 2:
        ws.delete_rows(2, max_r)

# 4. Clean Gantt
if "Gantt" in wb.sheetnames:
    ws = wb["Gantt"]
    for row in ws.iter_rows(min_row=5, max_row=100, min_col=1, max_col=1):
        if row[0].value:
            row[0].value = None

# 5. Clean Bitácora completely
ws = wb["Bitácora"]

# Clear OG
for row in ws.iter_rows(min_row=1, max_row=10):
    if row[1].value == "OG":
        row[2].value = "" # Clear Description

# Clear all OE descriptions (look for "Objetivo" and "Descripción" header)
reading_oe = False
for row in ws.iter_rows(min_row=1, max_row=50):
    col_b = row[1].value
    col_c = row[2].value
    
    if col_b == "Objetivo" and col_c == "Descripción":
        reading_oe = True
        continue
        
    if reading_oe:
        if not col_b and not col_c:
            reading_oe = False
        elif col_b and str(col_b).startswith("OE"):
            row[2].value = "" # Clear Description

wb.save("project_tracking.xlsx")
print("Fixed project_tracking.xlsx completely.")
